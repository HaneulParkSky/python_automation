
# 업무 자동화를 위한 엑셀 자동화

# 엑셀 파일 읽기
# openpyxl 라이브러리 활용 : xlsx 파일 읽고, 저장 모두 가능

import openpyxl

# 엑셀 파일 오픈 (openpyxl.load_workbook())

excel_file = openpyxl.load_workbook('shop_in_seoul.xlsx')

# 엑셀 파일 안에 시트 이름 확인 (excel_file.sheetnames)

excel_sheet = excel_file["fastfood"]

# 시트 안에 있는 데이터 읽기
# item 에는 한 라인의 각 셀에 있는 데이터를 가져옴
# 각 데이터는 각 리스트 아이템의 value 변수로부터 실제 데이터를 가져올 수 있음

for item in excel_sheet.rows:
    print(item[0].value, item[1].value, item[2].value)

# 오픈한 엑셀 파일 닫기
excel_file.close()
